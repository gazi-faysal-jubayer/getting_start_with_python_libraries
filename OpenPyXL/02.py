import openpyxl

# iter_row
workbook = openpyxl.load_workbook("OpenPyXL/mydata.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row =2, max_row = 7, values_only = True):
    print(row)
    name, age, city = row
    print(name)
    print(age)
    print(city)
    
    """
    we can also use 'row_offset" and 'column_offset' as well
    main_row, max_row, min_col, max_col, row_offset and column_offset combined:
    
    for row in sheet.iter_rows(min_row =2, max_row = 7, min_col = 2, max_col = 4, row_offset=2, column_offset=2 values_only = True):
    """